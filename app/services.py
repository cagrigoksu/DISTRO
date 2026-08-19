import csv
import io
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

import config
from .db import get_db, get_or_create, get_or_create_brand, get_or_create_model

POST_COLUMNS = [
    "Internal_Number", "Name", "Surname", "CNS", "Nationality",
    "ONA_Number", "MNA", "Preference", "PRF"
]

def clean(v):
    if pd.isna(v):
        return ""
    return str(v).strip()

def read_csv_upload(file_storage):
    raw = file_storage.read()
    text = raw.decode("utf-8-sig")
    df = pd.read_csv(io.StringIO(text), dtype=str, keep_default_na=False)
    df.columns = [str(c).strip() for c in df.columns]
    missing = [c for c in POST_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError("Missing CSV columns: " + ", ".join(missing))
    return df[POST_COLUMNS].fillna("")

def recipient_upsert(conn, national_id, name="", surname="", cns="", nationality="", ona=""):
    row = conn.execute(
        "SELECT id FROM recipients WHERE national_id=?", (national_id,)
    ).fetchone()
    if row:
        conn.execute(
            """UPDATE recipients
               SET name=?, surname=?, cns=?, nationality=?, ona_number=?
               WHERE id=?""",
            (name, surname, cns, nationality, ona, row["id"])
        )
        return row["id"]
    conn.execute(
        """INSERT INTO recipients(national_id,name,surname,cns,nationality,ona_number)
           VALUES (?,?,?,?,?,?)""",
        (national_id, name, surname, cns, nationality, ona)
    )
    return conn.execute("SELECT last_insert_rowid()").fetchone()[0]

def post_distribute(df, distribution_date):
    conn = get_db()
    errors = []
    updated = []
    try:
        for idx, row in df.iterrows():
            internal = clean(row["Internal_Number"])
            if not internal:
                errors.append(f"CSV row {idx+2}: Internal_Number is blank.")
                continue
            device = conn.execute(
                """SELECT d.*, m.name AS model_name
                   FROM devices d
                   JOIN models m ON m.id=d.model_id
                   WHERE d.internal_barcode=?""", (internal,)
            ).fetchone()
            if not device:
                errors.append(f"Internal number '{internal}' was not found in the database.")
                continue

            national_id = clean(row["CNS"])
            if not national_id:
                errors.append(f"CSV row {idx+2}: CNS is blank for '{internal}'.")
                continue

            recipient_id = recipient_upsert(
                conn, national_id,
                clean(row["Name"]), clean(row["Surname"]), national_id,
                clean(row["Nationality"]), clean(row["ONA_Number"])
            )

            conn.execute(
                """UPDATE devices
                   SET recipient_id=?, distribution_date=?, is_distributed=1,
                       updated_at=CURRENT_TIMESTAMP
                   WHERE id=?""",
                (recipient_id, distribution_date, device["id"])
            )
            conn.execute(
                "INSERT INTO audit_log(action,entity,entity_id,details) VALUES (?,?,?,?)",
                ("POST_DISTRIBUTION", "device", device["id"],
                 f"recipient={national_id}; distribution_date={distribution_date}")
            )
            updated.append({
                "internal_number": internal,
                "recipient_id": national_id,
                "model": device["model_name"],
            })

        if errors:
            conn.rollback()
        else:
            conn.commit()
        return updated, errors
    finally:
        conn.close()

def update_sp_distribution_file(updated, distribution_date):
    source = Path(config.SP_DISTRIBUTION_FILE)
    if not source.exists():
        raise FileNotFoundError(f"SP Distribution file not found: {source}")

    wb = load_workbook(source)
    if config.SP_DISTRIBUTION_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{config.SP_DISTRIBUTION_SHEET}' not found in {source.name}"
        )
    ws = wb[config.SP_DISTRIBUTION_SHEET]

    # column O = 15, AE = 31, AG = 33, AI = 35, AR = 44
    recipient_to_row = {}
    for r in range(1, ws.max_row + 1):
        value = ws.cell(r, 15).value
        if value is not None:
            recipient_to_row[str(value).strip()] = r

    not_found = []
    for item in updated:
        recipient = str(item["recipient_id"]).strip()
        r = recipient_to_row.get(recipient)
        if not r:
            not_found.append(recipient)
            continue
        ws.cell(r, 31).value = distribution_date
        ws.cell(r, 33).value = item["internal_number"]
        ws.cell(r, 35).value = item["model"]
        ws.cell(r, 44).value = "Distributed"

    if not_found:
        raise ValueError(
            "These recipient IDs were not found in column O of the SP Distribution file: "
            + ", ".join(not_found)
        )
    wb.save(source)
    return source

def build_inventory_workbook(output_path):
    conn = get_db()
    try:
        devices = conn.execute(
            """SELECT
                d.id, d.internal_barcode, dt.name AS device_type,
                b.name AS brand, m.name AS model, c.name AS connection,
                d.is_engraved, d.engraving_date, e.name AS engraver,
                d.is_distributed, r.national_id AS recipient_id,
                d.distribution_date, s.name AS status, d.place AS place,
                d.capacity_gb AS capacity, d.os AS os,
                d.serial_number, d.imei_1, d.imei_2,
                donor_type.name AS donor_type, d.entry_date, donor.name AS donor
            FROM devices d
            JOIN device_types dt ON dt.id=d.device_type_id
            JOIN brands b ON b.id=d.brand_id
            JOIN models m ON m.id=d.model_id
            JOIN connections c ON c.id=d.connection_id
            LEFT JOIN engravers e ON e.id=d.engraver_id
            LEFT JOIN recipients r ON r.id=d.recipient_id
            JOIN statuses s ON s.id=d.status_id
            JOIN donor_types donor_type ON donor_type.id=d.donor_type_id
            LEFT JOIN donors donor ON donor.id=d.donor_id
            ORDER BY d.id"""
        ).fetchall()

        accessories = conn.execute(
            "SELECT name, quantity FROM accessories ORDER BY name"
        ).fetchall()

        stock_counts = []
        for label, dtype, apple in [
            ("SP Android", "Phone", False),
            ("SP iPhone", "Phone", True),
            ("TAB Android", "Tablet", False),
            ("Tab iPad", "Tablet", True),
        ]:
            op = "=" if apple else "!="
            rfd = conn.execute(
                f"""SELECT COUNT(*) FROM devices d
                    JOIN device_types dt ON dt.id=d.device_type_id
                    JOIN brands b ON b.id=d.brand_id
                    WHERE dt.name=? AND lower(b.name) {op} lower('Apple')
                      AND d.is_engraved=1 AND d.is_distributed=0""",
                (dtype,)
            ).fetchone()[0]
            pending = conn.execute(
                f"""SELECT COUNT(*) FROM devices d
                    JOIN device_types dt ON dt.id=d.device_type_id
                    JOIN brands b ON b.id=d.brand_id
                    WHERE dt.name=? AND lower(b.name) {op} lower('Apple')
                      AND d.is_engraved=0 AND d.is_distributed=0""",
                (dtype,)
            ).fetchone()[0]
            stock_counts.append((label, rfd, pending))
    finally:
        conn.close()

    wb = load_workbook(config.BASE_DIR / "templates" / "inventory_base.xlsx")
    ws_analysis = wb["Stock Analysis"]
    ws_device = wb["Device"]
    ws_acc = wb["Accessoires"]

    # clear existing data areas
    for ws in [ws_analysis, ws_device, ws_acc]:
        for row in ws.iter_rows():
            for cell in row:
                if cell.row > 1:
                    cell.value = None

    ws_acc.append(["Accessoire", "Number"])
    for a in accessories:
        ws_acc.append([a["name"], a["quantity"]])

    headers = [
        "id", "internal_barcode", "device type", "brand", "model", "connection",
        "is_engraved", "engraving_date", "engraver", "is_distributed",
        "recipient_id", "distribution_date", "status", "place", "capacity",
        "OS", "serial_number", "imei-1", "imei-2", "donor_type", "entry_date", "donor"
    ]
    ws_device.append(headers)
    for d in devices:
        ws_device.append([d[h] for h in [
            "id", "internal_barcode", "device_type", "brand", "model", "connection",
            "is_engraved", "engraving_date", "engraver", "is_distributed",
            "recipient_id", "distribution_date", "status", "place", "capacity",
            "os", "serial_number", "imei_1", "imei_2", "donor_type", "entry_date", "donor"
        ]])

    ws_analysis["A1"] = "RFD Stock"
    ws_analysis["A2"], ws_analysis["B2"], ws_analysis["C2"] = "Device Type", "Stock", "Location"
    ws_analysis["E1"] = "Pending Refurbishment"
    ws_analysis["E2"], ws_analysis["F2"], ws_analysis["G2"] = "Device Type", "Stock", "Location"

    # Apple vs non-Apple
    for i, (label, count_rfd, count_pending) in enumerate(stock_counts, start=3):
        ws_analysis.cell(i, 1).value = label
        ws_analysis.cell(i, 2).value = count_rfd
        ws_analysis.cell(i, 3).value = "Office Cabinet 12"
        ws_analysis.cell(i, 5).value = label
        ws_analysis.cell(i, 6).value = count_pending
        ws_analysis.cell(i, 7).value = "Office Cabinet 12"

    # formatting
    for ws in [ws_analysis, ws_device, ws_acc]:
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 35)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path

def make_inventory_template():
    from openpyxl import Workbook
    path = config.BASE_DIR / "templates" / "inventory_base.xlsx"
    if path.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Analysis"
    wb.create_sheet("Device")
    wb.create_sheet("Accessoires")
    wb.save(path)

def generate_pre_distribution(df, output_name):
    template = Path(config.PRE_DISTRIBUTION_TEMPLATE)
    if not template.exists():
        raise FileNotFoundError(f"Pre-Distribution template not found: {template}")

    safe_name = Path(output_name).name
    if not safe_name.lower().endswith(".xlsx"):
        safe_name += ".xlsx"
    destination = Path(config.PRE_DISTRIBUTION_OUTPUT_DIR) / safe_name
    shutil.copy2(template, destination)

    wb = load_workbook(destination)
    required_sheets = ["Reception List", "contract template", "registration form"]
    missing = [s for s in required_sheets if s not in wb.sheetnames]
    if missing:
        raise ValueError("Template is missing sheets: " + ", ".join(missing))

    # Reception List: D,E,F,G,H,I,J,K,M from row 4
    ws = wb["Reception List"]
    for i, row in df.iterrows():
        r = 4 + i
        ws.cell(r, 4).value = clean(row["Internal_Number"])
        ws.cell(r, 5).value = clean(row["Name"])
        ws.cell(r, 6).value = clean(row["Surname"])
        ws.cell(r, 7).value = clean(row["CNS"])
        ws.cell(r, 8).value = clean(row["Nationality"])
        ws.cell(r, 9).value = clean(row["ONA_Number"])
        ws.cell(r, 10).value = clean(row["MNA"]).upper() if clean(row["MNA"]) else ""
        ws.cell(r, 11).value = clean(row["Preference"])
        ws.cell(r, 13).value = clean(row["PRF"]).upper() if clean(row["PRF"]) else ""

    # contract template: B,C,D,E,F,G,I,J from row 2
    ws = wb["contract template"]
    for i, row in df.iterrows():
        r = 2 + i
        ws.cell(r, 2).value = clean(row["Internal_Number"])
        ws.cell(r, 3).value = clean(row["Name"])
        ws.cell(r, 4).value = clean(row["Surname"])
        ws.cell(r, 5).value = clean(row["CNS"])
        ws.cell(r, 6).value = clean(row["Nationality"])
        ws.cell(r, 7).value = clean(row["ONA_Number"])
        ws.cell(r, 9).value = clean(row["MNA"]).upper() if clean(row["MNA"]) else ""
        ws.cell(r, 10).value = clean(row["PRF"]).upper() if clean(row["PRF"]) else ""

    # registration form: B,A,C,D,E,K from row 2
    ws = wb["registration form"]
    for i, row in df.iterrows():
        r = 2 + i
        ws.cell(r, 2).value = clean(row["Name"])
        ws.cell(r, 1).value = clean(row["Surname"])
        ws.cell(r, 3).value = clean(row["CNS"])
        ws.cell(r, 4).value = clean(row["Nationality"])
        ws.cell(r, 5).value = clean(row["ONA_Number"])
        ws.cell(r, 11).value = clean(row["PRF"]).upper() if clean(row["PRF"]) else ""

    wb.save(destination)
    return destination
